import csv
import json
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config.config import Settings
from evaluation.dataset import DatasetSample, load_dataset
from evaluation.metrics import aggregate, exact_match, token_scores
from react.agent import ReActAgent

try:
    import matplotlib.pyplot as plt
except Exception:  # pragma: no cover - matplotlib is optional at import time
    plt = None


COMPARISON_COLUMNS = [
    "Model",
    "Dataset",
    "System",
    "n",
    "Accuracy",
    "Precision",
    "Recall",
    "Token F1",
    "Sem F1",
    "MAR",
    "MKG Matches / n",
    "Retrieval Failures",
    "Reasoning Failures",
    "Accuracy Improvement",
    "Sem F1 Change",
    "Average Latency",
    "Average Iterations",
    "Notes",
]

PER_QUESTION_COLUMNS = [
    "Model",
    "Dataset",
    "System",
    "n",
    "Question",
    "Prediction",
    "Expected Answer",
    "Exact Match",
    "Precision",
    "Recall",
    "F1",
    "Latency",
    "Iterations",
    "Trace Path",
]

GRAPH_METRICS = [
    ("accuracy", "Accuracy", "accuracy.png"),
    ("precision", "Precision", "precision.png"),
    ("recall", "Recall", "recall.png"),
    ("f1", "Token F1", "token_f1.png"),
    ("average_latency", "Average Latency (s)", "latency.png"),
    ("average_iterations", "Average Iterations", "iterations.png"),
    ("retrieval_failures", "Retrieval Failures", "retrieval_failures.png"),
    ("reasoning_failures", "Reasoning Failures", "reasoning_failures.png"),
]


@dataclass(frozen=True)
class ComparisonRun:
    model: str
    dataset_name: str
    system_name: str
    sample_size: int
    metrics: dict[str, float]
    retrieval_failures: int
    reasoning_failures: int
    notes: str

    def row(self) -> dict[str, str]:
        return {
            "Model": self.model,
            "Dataset": self.dataset_name,
            "System": self.system_name,
            "n": str(self.sample_size),
            "Accuracy": f"{self.metrics['accuracy']:.3f}",
            "Precision": f"{self.metrics['precision']:.3f}",
            "Recall": f"{self.metrics['recall']:.3f}",
            "Token F1": f"{self.metrics['f1']:.3f}",
            "Sem F1": "N/A",
            "MAR": "N/A",
            "MKG Matches / n": "N/A",
            "Retrieval Failures": str(self.retrieval_failures),
            "Reasoning Failures": str(self.reasoning_failures),
            "Accuracy Improvement": "N/A",
            "Sem F1 Change": "N/A",
            "Average Latency": f"{self.metrics['average_latency']:.2f}",
            "Average Iterations": f"{self.metrics['average_iterations']:.2f}",
            "Notes": self.notes,
        }


@dataclass(frozen=True)
class ExperimentPaths:
    root: Path
    graphs: Path
    traces: Path


def compare_models(
    agent_factory,
    settings: Settings,
    dataset_path: Path,
    output_path: Path,
    models: list[str],
    sample_sizes: list[int],
    dataset_name: str,
    system_name: str = "ReAct Baseline",
) -> dict[str, object]:
    dataset = load_dataset(dataset_path)
    paths = _create_experiment_paths(output_path)
    runs: list[ComparisonRun] = []
    per_question_rows: list[dict[str, str]] = []
    summary_details: list[dict[str, object]] = []

    for sample_size in sample_sizes:
        samples = dataset[:sample_size]
        if len(samples) < sample_size:
            raise ValueError(
                f"Requested sample size n={sample_size}, but dataset only has {len(dataset)} samples."
            )

        for model in models:
            settings.model_name = model
            agent = agent_factory(settings)
            run_results = _evaluate_samples(
                agent=agent,
                samples=samples,
                model=model,
                dataset_name=dataset_name,
                system_name=system_name,
                sample_size=sample_size,
                trace_dir=paths.traces,
            )
            metrics = aggregate(run_results)
            retrieval_failures, reasoning_failures = _classify_failures(run_results)
            notes = "Ollama local baseline; non-ReAct columns are N/A"
            run = ComparisonRun(
                model=model,
                dataset_name=dataset_name,
                system_name=system_name,
                sample_size=sample_size,
                metrics=metrics.__dict__,
                retrieval_failures=retrieval_failures,
                reasoning_failures=reasoning_failures,
                notes=notes,
            )
            runs.append(run)
            per_question_rows.extend(_per_question_rows(run_results, model, dataset_name, system_name, sample_size))
            summary_details.append(
                {
                    "model": model,
                    "dataset": dataset_name,
                    "system": system_name,
                    "sample_size": sample_size,
                    "metrics": metrics.__dict__,
                    "retrieval_failures": retrieval_failures,
                    "reasoning_failures": reasoning_failures,
                    "trace_count": len(run_results),
                }
            )

    payload = {
        "experiment_dir": str(paths.root),
        "columns": COMPARISON_COLUMNS,
        "runs": [run.row() for run in runs],
        "details": summary_details,
    }
    _write_outputs(payload, per_question_rows, paths)
    return payload


def _evaluate_samples(
    agent: ReActAgent,
    samples: list[DatasetSample],
    model: str,
    dataset_name: str,
    system_name: str,
    sample_size: int,
    trace_dir: Path,
) -> list[dict[str, object]]:
    results: list[dict[str, object]] = []
    for index, sample in enumerate(samples, start=1):
        started = time.perf_counter()
        state = agent.run(sample.question)
        latency = time.perf_counter() - started
        prediction = state.final_answer or ""
        precision, recall, f1 = token_scores(prediction, sample.answer)
        trace_path = _write_question_trace(
            trace_dir=trace_dir,
            model=model,
            dataset_name=dataset_name,
            sample_size=sample_size,
            index=index,
            question=sample.question,
            expected_answer=sample.answer,
            prediction=prediction,
            latency=latency,
            state=state,
        )
        results.append(
            {
                "question": sample.question,
                "expected_answer": sample.answer,
                "prediction": prediction,
                "exact_match": exact_match(prediction, sample.answer),
                "precision": precision,
                "recall": recall,
                "f1": f1,
                "latency_seconds": latency,
                "iterations": state.iteration_count,
                "trace_path": str(trace_path),
                "trajectory": state.trajectory(),
            }
        )
    return results


def _write_question_trace(
    trace_dir: Path,
    model: str,
    dataset_name: str,
    sample_size: int,
    index: int,
    question: str,
    expected_answer: str,
    prediction: str,
    latency: float,
    state,
) -> Path:
    safe_model = _safe_name(model)
    path = trace_dir / f"{dataset_name}_n{sample_size}_{safe_model}_{index:04d}.json"
    payload = {
        "model": model,
        "dataset": dataset_name,
        "sample_size": sample_size,
        "question_index": index,
        "question": question,
        "expected_answer": expected_answer,
        "final_answer": prediction,
        "iterations": state.iteration_count,
        "latency_seconds": latency,
        "trajectory": state.trajectory(),
    }
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return path


def _classify_failures(results: list[dict[str, object]]) -> tuple[int, int]:
    retrieval_failures = 0
    reasoning_failures = 0

    for result in results:
        if result["exact_match"]:
            continue
        trajectory = result.get("trajectory", [])
        observations = [
            str(step.get("observation", ""))
            for step in trajectory
            if isinstance(step, dict) and step.get("observation")
        ]
        if any("No matching documents found." in observation for observation in observations):
            retrieval_failures += 1
        elif observations:
            reasoning_failures += 1

    return retrieval_failures, reasoning_failures


def _per_question_rows(
    results: list[dict[str, object]],
    model: str,
    dataset_name: str,
    system_name: str,
    sample_size: int,
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for result in results:
        rows.append(
            {
                "Model": model,
                "Dataset": dataset_name,
                "System": system_name,
                "n": str(sample_size),
                "Question": str(result["question"]),
                "Prediction": str(result["prediction"]),
                "Expected Answer": str(result["expected_answer"]),
                "Exact Match": str(result["exact_match"]),
                "Precision": f"{float(result['precision']):.3f}",
                "Recall": f"{float(result['recall']):.3f}",
                "F1": f"{float(result['f1']):.3f}",
                "Latency": f"{float(result['latency_seconds']):.2f}",
                "Iterations": str(result["iterations"]),
                "Trace Path": str(result["trace_path"]),
            }
        )
    return rows


def _create_experiment_paths(output_path: Path) -> ExperimentPaths:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    root = output_path.parent / f"{output_path.stem}_{timestamp}"
    graphs = root / "graphs"
    traces = root / "logs" / "traces"
    graphs.mkdir(parents=True, exist_ok=True)
    traces.mkdir(parents=True, exist_ok=True)
    return ExperimentPaths(root=root, graphs=graphs, traces=traces)


def _write_outputs(payload: dict[str, object], per_question_rows: list[dict[str, str]], paths: ExperimentPaths) -> None:
    runs = list(payload["runs"])
    _write_txt(paths.root / "comparison.txt", runs)
    _write_csv(paths.root / "comparison.csv", runs)
    _write_json(paths.root / "comparison.json", payload)
    _write_workbook(paths.root / "comparison.xlsx", runs, per_question_rows)
    _write_graphs(paths.graphs, runs)


def _write_txt(path: Path, runs: list[dict[str, str]]) -> None:
    path.write_text(format_comparison_table(runs), encoding="utf-8")


def _write_csv(path: Path, runs: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=COMPARISON_COLUMNS)
        writer.writeheader()
        writer.writerows(runs)


def _write_json(path: Path, payload: dict[str, object]) -> None:
    summary = {
        "columns": payload["columns"],
        "runs": payload["runs"],
        "details": payload["details"],
    }
    path.write_text(json.dumps(summary, indent=2), encoding="utf-8")


def _write_workbook(path: Path, runs: list[dict[str, str]], per_question_rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    comparison_sheet = workbook.active
    comparison_sheet.title = "Model comparison"
    _write_sheet(comparison_sheet, COMPARISON_COLUMNS, runs)
    question_sheet = workbook.create_sheet("Per-question evaluation")
    _write_sheet(question_sheet, PER_QUESTION_COLUMNS, per_question_rows)
    workbook.save(path)


def _write_sheet(sheet, columns: list[str], rows: list[dict[str, str]]) -> None:
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])
    sheet.freeze_panes = "A2"
    for index, column in enumerate(columns, start=1):
        width = min(max(len(column), *(len(str(row.get(column, ""))) for row in rows), 12), 60)
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_graphs(graph_dir: Path, runs: list[dict[str, str]]) -> None:
    if plt is None or not runs:
        return
    models = sorted({row["Model"] for row in runs})
    sample_sizes = sorted({row["n"] for row in runs}, key=lambda value: int(value))
    colors = plt.get_cmap("tab10")

    for key, title, filename in GRAPH_METRICS:
        fig, ax = plt.subplots(figsize=(10, 6), dpi=200)
        x_positions = list(range(len(sample_sizes)))
        width = 0.8 / max(len(models), 1)
        for model_index, model in enumerate(models):
            values = []
            for sample_size in sample_sizes:
                row = next((item for item in runs if item["Model"] == model and item["n"] == sample_size), None)
                values.append(_metric_value(row, key) if row is not None else 0.0)
            offsets = [x + (model_index - (len(models) - 1) / 2) * width for x in x_positions]
            ax.bar(offsets, values, width=width, label=model, color=colors(model_index % 10))
        ax.set_title(title)
        ax.set_xlabel("Sample size (n)")
        ax.set_ylabel(title)
        ax.set_xticks(x_positions)
        ax.set_xticklabels(sample_sizes)
        ax.legend()
        ax.grid(axis="y", linestyle="--", alpha=0.35)
        fig.tight_layout()
        fig.savefig(graph_dir / filename)
        plt.close(fig)


def _metric_value(row: dict[str, str], key: str) -> float:
    mapping = {
        "accuracy": "Accuracy",
        "precision": "Precision",
        "recall": "Recall",
        "f1": "Token F1",
        "average_latency": "Average Latency",
        "average_iterations": "Average Iterations",
        "retrieval_failures": "Retrieval Failures",
        "reasoning_failures": "Reasoning Failures",
    }
    value = row[mapping[key]].replace("s", "")
    return float(value)


def format_comparison_table(rows: object) -> str:
    typed_rows = list(rows)
    if not typed_rows:
        return "No comparison runs were produced.\n"

    widths = {
        column: max(len(column), *(len(str(row[column])) for row in typed_rows))
        for column in COMPARISON_COLUMNS
    }
    separator = " | "
    header = separator.join(column.ljust(widths[column]) for column in COMPARISON_COLUMNS)
    divider = "-+-".join("-" * widths[column] for column in COMPARISON_COLUMNS)
    body = [
        separator.join(str(row[column]).ljust(widths[column]) for column in COMPARISON_COLUMNS)
        for row in typed_rows
    ]
    return "\n".join([header, divider, *body]) + "\n"


def _safe_name(value: str) -> str:
    return "".join(char if char.isalnum() else "_" for char in value).strip("_")
